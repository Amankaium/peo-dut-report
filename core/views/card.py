from django.views import View
from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework import generics
from ..models import *
from ..serializers import *
from django.contrib import messages
from openpyxl import load_workbook


class CardAddView(View):
    def get(self, request):
        return render(request, 'core/cards_add.html')

    def post(self, request):
        excel_file = request.FILES["excel_file"]
        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")
        context = {}
        excel_file_source = load_workbook('media/' + new_excel_source.excel_file.name)
        page = excel_file_source[excel_file_source.sheetnames[0]]
        created_qty = 0
        for row in page:
            if row[0].row == 1:
                continue
            number = row[0].value
            new_card, created = Card.objects.get_or_create(number=number, id_realcom=1)
            if created:
                created_qty += 1
        messages.success(request, f"Добавлено {created_qty} карт(ы)")
        return render(request, 'core/cards_add.html', context)



class CardListCreateAPIView(generics.ListCreateAPIView):
    queryset = Card.objects.all()
    serializer_class = CardSerializer

class CardDetailAPIView(RetrieveAPIView):
    queryset = Card.objects.all()
    serializer_class = CardSerializer


class CardListView(View):
    def get(self, request):
        context = {}
        context["cards"] = Card.objects.all()
        return render(request, 'core/cards.html', context)



