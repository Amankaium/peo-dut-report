from datetime import datetime
import calendar
import locale
from openpyxl import load_workbook
from django.views import View
from django.views.generic.edit import CreateView
from django.shortcuts import render, redirect, HttpResponse
from django.contrib import messages
from rest_framework.viewsets import ModelViewSet
from core.serializers import *
from core.models import DeltaReport, MonthReport
import requests


locale.setlocale(locale.LC_TIME, 'ru_RU')
month_names = [month for month in calendar.month_name]


class DeltaReportAddView(CreateView):
    model = MonthReport
    fields = ['month', 'year']
    template_name = "core/delta_reports_add.html"

    def post(self, request):
        context = {}

        excel_file = request.FILES["excel_file"]
        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")

        # try:
        #     new_month_report = MonthReport.objects.create(
        #         source=new_excel_source,
        #         month=request.POST["month"],
        #         year=request.POST["year"]
        #     )
        # except:
        #     return HttpResponse(f"Ошибка! Такой отчёт уже существует")

        month, year = request.POST["month"], request.POST["year"]

        if not MonthReport.objects.filter(month=month, year=year):
            new_month_report = MonthReport.objects.create(
                source=new_excel_source,
                month=request.POST["month"],
                year=request.POST["year"]
            )
        else:
            new_month_report = MonthReport.objects.get(
                month=request.POST["month"],
                year=request.POST["year"]
            )
            new_month_report.source = new_excel_source
            new_month_report.save()
        
            
        messages.success(request, f"Создан {new_month_report}")

        excel_file_source = load_workbook('media/' + new_excel_source.excel_file.name)
        # page = excel_file_source["Проверка ПЭО"]
        page = excel_file_source[excel_file_source.sheetnames[0]]
        delta_reports_to_create = []

        for row in page:
            if row[0].row == 1:
                continue
            
            try:
                car_name = row[1].value

                if not Transport.objects.filter(name=car_name).exists():
                    print(car_name)
                    continue
                if car_name == "нет ДУТ":
                    continue
                if car_name is None or car_name in ['', ' ']:
                    break

                car_object, car_created = Transport.objects.get_or_create(
                    name=car_name
                )

                date_format = "%Y %B %d %H:%M"
                period_start = datetime.strptime(row[2].value, date_format)
                period_end = datetime.strptime(row[3].value, date_format)

                fact_km = row[4].value
                start_balance = row[7].value
                fueling_gsm = row[9].value
                actual_fuel_consumption = row[11].value
                norm_fuel_consumption = row[12].value
                departure = row[13].value
                actual = row[16].value
                fuel_calculation_norm = row[17].value
                departure_balance = row[18].value
                end_balance = row[19].value
                end_mech_balance = row[21].value
                difference = row[22].value
                deficiency = row[23].value
                note = row[25].value

                new_delta_report = DeltaReport(
                    month_report=new_month_report,
                    transport=car_object,
                    vehicle_name=car_name,
                    period_start=period_start,
                    period_end=period_end,
                    fact_km=fact_km,
                    start_balance=start_balance,
                    fueling_gsm=fueling_gsm,
                    actual_fuel_consumption=actual_fuel_consumption,
                    norm_fuel_consumption=norm_fuel_consumption,
                    departure=departure,
                    actual=actual,
                    fuel_calculation_norm=fuel_calculation_norm,
                    departure_balance=departure_balance,
                    end_balance=end_balance,
                    end_mech_balance=end_mech_balance,
                    difference=difference,
                    deficiency=deficiency,
                    note=note,
                )
                delta_reports_to_create.append(new_delta_report)
            except BaseException as e:
                print(f"{car_name}\t{e}\t")
        DeltaReport.objects.bulk_create(delta_reports_to_create)
           
        messages.success(request, f"Добавлен отчёт об отклонениях")
        return redirect('home')


class DeltaReportUpdateView(View):
    template_name = 'core/delta_reports_update.html'

    def get_context(self):
        month_reports = MonthReport.objects.all()
        # q = len(month_reports)
        # if q > 3:
        #     month_reports = month_reports[q-3:]
        month_report = MonthReport.objects.get(id=self.kwargs["id"])
        delta_reports = DeltaReport.objects.filter(month_report=month_report)
        context = {
            "month_reports": month_reports,
            "month_report": month_report,
            "delta_reports": delta_reports,
        }
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self.get_context())

    def post(self, request, *args, **kwargs):
        context = self.get_context()
        sid = "9b0fd860f877c51459b7b8275150ca1c"
        for obj in context["delta_reports"]:
            # try:
                period_start = int(obj.period_start.timestamp())
                period_end = int(obj.period_end.timestamp())
                id_realcom = obj.transport.id_realcom

                response = requests.post(
                    url=f'http://wialon.realcom.kg/wialon/ajax.html?svc=report/exec_report&sid={sid}',
                    data={
                        "sid": sid,
                        "params": '{"reportResourceId":6354,"reportTemplateId":3,"reportTemplate":null,"reportObjectId":%s,"reportObjectSecId":0,"interval":{"flags":16777216,"from":%s,"to":%s},"remoteExec":0}' % (id_realcom, period_start, period_end)
                    }
                )
                # print(id_realcom, period_start, period_end)
                data = response.json()
                # print(data)
                # odometer_mileage = data["reportResult"]["stats"][6][1].split()[0]
                # trip_mileage = data["reportResult"]["stats"][7][1].split()[0]
                # start_level = data["reportResult"]["stats"][16][1].split()[0]
                # total_refueled = data["reportResult"]["stats"][19][1].split()[0]
                # avg_trip_dut_consumption = data["reportResult"]["stats"][14][1].split()[0]
                # avg_dut_consumption = data["reportResult"]["stats"][15][1].split()[0]
                # end_level = data["reportResult"]["stats"][17][1].split()[0]
                # difference = data["reportResult"]["stats"][22][1].split()[0]

                # Объект
                # Время выполнения отчета
                # Начало интервала                  # period_start
                # Окончание интервала               # period_end
                # Время в движении
                # Пробег в поездках                 # trip_mileage
                # Потрачено по ДУТ
                # Ср. расход по ДУТ в поездках      # avg_trip_dut_consumption
                # Ср. расход по ДУТ (весь пробег)   # avg_dut_consumption
                # Нач. уровень",                    # start_level
                # Кон. уровень                      # end_level
                # Всего топлива слито               # total_fuel_drained
                # Всего заправлено                  # total_refueled
                # Разница                           # difference

                value_texts = {
                    'period_start': "Начало интервала",
                    'period_end': "Окончание интервала",
                    'trip_mileage': "Пробег в поездках",
                    'avg_trip_dut_consumption': "Ср. расход по ДУТ в поездках",
                    'avg_dut_consumption': "Ср. расход по ДУТ (весь пробег)",
                    'start_level': "Нач. уровень",
                    'end_level': "Кон. уровень",
                    'total_fuel_drained': "Всего топлива слито",
                    'total_refueled': "Всего заправлено",
                    'difference': "Разница",
                }

                stats = data["reportResult"]["stats"]
                stats_index = {}
                for row, i in enumerate(stats):
                    txt = row[0].split()[0]
                    stats_index[txt] = i

                # todo via setattr
                trip_mileage = stats[stats_index[value_texts["trip_mileage"]]].split()[0]
                avg_trip_dut_consumption = stats[stats_index[value_texts["avg_trip_dut_consumption"]]].split()[0]
                avg_dut_consumption = stats[stats_index[value_texts["avg_dut_consumption"]]].split()[0]
                start_level = stats[stats_index[value_texts["start_level"]]].split()[0]
                end_level = stats[stats_index[value_texts["end_level"]]].split()[0]
                total_fuel_drained = stats[stats_index[value_texts["total_fuel_drained"]]].split()[0]
                total_refueled = stats[stats_index[value_texts["total_refueled"]]].split()[0]
                difference = stats[stats_index[value_texts["difference"]]].split()[0]

                not_valid_val = ['-----', '', ' ']

                # obj.odometer_mileage = float(odometer_mileage) if odometer_mileage not in not_valid_val else None
                obj.trip_mileage = float(trip_mileage) if trip_mileage not in not_valid_val else None
                obj.start_level = float(start_level) if start_level not in not_valid_val else None
                obj.total_refueled = float(total_refueled) if total_refueled not in not_valid_val else None
                obj.avg_trip_dut_consumption = float(avg_trip_dut_consumption) if avg_trip_dut_consumption not in not_valid_val else None
                obj.avg_dut_consumption = float(avg_dut_consumption) if avg_dut_consumption not in not_valid_val else None
                obj.end_level = float(end_level) if end_level not in not_valid_val else None
                obj.difference = float(total_fuel_drained) if total_fuel_drained not in not_valid_val else None
                obj.difference = float(difference) if difference not in not_valid_val else None

                obj.save()
                print(f"saved for {obj.transport.name}")
            # except BaseException as e:
            #     print(obj.transport.name, e)
            #     break
        context["data"] = data
        return render(request, self.template_name, context)


class DeltaReportViewSet(ModelViewSet):
    queryset = DeltaReport.objects.all()
    serializer_class = DeltaReportSerializer
