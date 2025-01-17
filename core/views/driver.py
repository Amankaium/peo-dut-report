from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework import generics
from rest_framework.viewsets import ModelViewSet
from core.serializers import *
from django.views import View
from django.shortcuts import render
from django.contrib import messages
from core.models import *
from openpyxl import load_workbook
from django.http import Http404


class DriversAddView(View):
    def get(self, request):
        return render(request, 'core/drivers_add.html')
    
    def post(self, request):
        excel_file = request.FILES["excel_file"]
        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")

        context = {}
        excel_file_source = load_workbook(new_excel_source.excel_file.url[1:])
        page = excel_file_source[excel_file_source.sheetnames[0]]
        # drivers_to_create = []
        created_qty = 0
        for row in page:
            if row[0].row == 1:
                continue

            full_name = row[2].value
            new_driver, created = DriversName.objects.get_or_create(full_name=full_name)
            if created:
                created_qty += 1
        #     new_driver = DriversName(
        #         full_name=full_name
        #     )
        #     drivers_to_create.append(new_driver)
        # DriversName.objects.bulk_create(drivers_to_create)
        # messages.success(request, f"Добавлено {len(drivers_to_create)} водителей")
        messages.success(request, f"Добавлено {created_qty} водителей")
        return render(request, 'core/drivers_add.html', context)


class GetDriversListView(View):
    def get(self, request):
        context = {}
        context['drivers'] = DriversName.objects.all()
        return render(request, 'core/get_drivers.html', context)


class DriverViewSet(ModelViewSet):
    queryset = DriversName.objects.all()
    serializer_class = DriverSerializer

class DriverListCreateAPIView(APIView):
    def get(self, request):
        drivers = DriversName.objects.all()
        serializer = DriverSerializer(
            instance=drivers,
            many=True
        )
        data = serializer.data
        return Response(data)
    
    def post(self, request):
        serializer = DriverSerializer(data=request.data)
        if serializer.is_valid():
            new_driver = serializer.save()
            new_serializer = DriverSerializer(instance=new_driver)
            return Response(new_serializer.data, 201)
        
        return Response(serializer.errors, 400)


class DriverDetailAPIView(APIView):
    def get(self, request, *args, **kwargs):
        id = kwargs["pk"]
        report = DriversName.objects.get(id=id)
        serializer = DriverSerializer(instance=report)
        return Response(serializer.data)


class DriverDetailView(View):
    template_name = 'core/driver_detail.html'  # Create this template

    def get(self, request, pk):
        try:
            driver = DriversName.objects.get(pk=pk)
        except DriversName.DoesNotExist:
            raise Http404("Driver does not exist")

        context = {'driver': driver}
        return render(request, self.template_name, context)


class DriversIdRealcomView(View):
    def get(self, request):
        return render(request, 'core/id_realcom_drivers_add.html')
    def post(self, request):
        excel_file = request.FILES["excel_file"]
        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")

        context = {}
        excel_file_source = load_workbook(new_excel_source.excel_file.url[1:])
        page = excel_file_source[excel_file_source.sheetnames[0]]
        created_qty = 0
        for row in page:
            if row[0].row == 1:
                continue
            id_realcom = row[0].value
            new_id_realcom, created = DriversName.objects.get_or_create(id_realcom=id_realcom)
            if created:
                created_qty += 1
            print(f'id_realcom: {id_realcom}, created: {created}')
        messages.success(request, f"Добавлено {created_qty} записей")
        return render(request, 'core/id_realcom_drivers_add.html', context)

class DriversRealcomView(View):
    def get(self, request):
        context = {}
        drivers = DriversName.objects.all()
        context['drivers'] = drivers
        return render(request, 'core/id_realcom_drivers_list.html', context)
