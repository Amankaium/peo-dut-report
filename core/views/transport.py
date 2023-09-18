from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from rest_framework.generics import ListAPIView, RetrieveAPIView, CreateAPIView
from rest_framework import generics
from core.serializers import *
from django.shortcuts import render
from django.views import View
from django.shortcuts import render
from core.models import *
from openpyxl import load_workbook
from django.contrib import messages

class TransportAddView(View):
    def get(self, request):
        return render(request, 'core/transport_add.html')
    def post(self, request):
        excel_file = request.FILES['excel_file']
        new_excel_source = ExcelSource.objects.create(excel_file=excel_file, created_by=request.user)
        messages.success(request, "Файл добавлен")
        context = {}
        excel_file_source = load_workbook(new_excel_source.excel_file.url[1:])
        page = excel_file_source[excel_file_source.sheetnames[0]]
        created_qty = 0
        for row in page:
            if row[0].row == 1:
                continue
            mark = row[0].value
            number = row[1].value
            name_realcom = row[2].value
            new_transport, created = Transport.objects.get_or_create(
                mark=mark,
                number=number,
                name_realcom=name_realcom,
            )
            if created:
                created_qty += 1
        messages.success(request, f'Добавлено {created_qty} транспортов')
        return render(request, 'core/transport_add.html', context)

class TransportInfoView(View):
    def get(self,request, *args, **kwargs):
        id = kwargs["pk"]
        context = {}
        context['transport'] = Transport.objects.get(id=id)
        return render(request, 'core/transports.html', context)
      


class GetTransportView(View):
    template_name = 'core/get_transports.html'
    def get(self, request):
        context = {}
        context['transports'] = Transport.objects.all()
        return render(request, self.template_name, context)

class TransportViewSet(ModelViewSet):
    queryset = Transport.objects.all()
    serializer_class = TransportSerializer

class TransportListAPIView(APIView):
    def get(self, request):
        transports = Transport.objects.all()
        serializer = TransportSerializer(
            instance=transports,
            many=True
        )
        data = serializer.data
        return Response(data)

class TransportDetailAPIView(APIView):
    def get(self, request, *args, **kwargs):
        id = kwargs["pk"]
        transport_object = Transport.objects.get(id=id)
        serializer = TransportSerializer(instance=transport_object)
        return Response(serializer.data)


class TransportCreateAPIView(APIView):
    def post(self, request, *args, **kwargs):
        data = request.data
        serializer = TransportSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                data={"message": "Успешно создано"},
                status=201,
            )
        
        return Response(
            data=serializer.errors,
            status=400
        )
