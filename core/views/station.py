from rest_framework.viewsets import ModelViewSet
from core.models import Station
from core.serializers import *
from django.views import View
from django.shortcuts import render
from django.contrib import messages
from openpyxl import load_workbook

class StationAddView(View):
    def get(self, request):
        return render(request, 'core/stations_add.html')

    def post(self, request):
        excel_file = request.FILES["excel_file"]

        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")
        context = {}
        excel_file_source = load_workbook('media/' + new_excel_source.excel_file.name)
        page = excel_file_source[excel_file_source.sheetnames[0]]
        created_qty = 0
        for row in page:
            if row[0].row == 1:
                continue
            name = row[0].value
            new_station, created = Station.objects.get_or_create(name=name, id_realcom=2)
            if created:
                created_qty += 1
        messages.success(request, f"Добавлено {created_qty} AЗС")
        return render(request, 'core/stations_add.html', context)

class StationViewSet(ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer

class StationListView(View):
    def get(self, request):
        context = {}
        context['stations'] = Station.objects.all()
        return render(request, 'core/station.html', context)


