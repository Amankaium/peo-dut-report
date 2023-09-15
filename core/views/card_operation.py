from core.serializers import *
from core.models import CardOperation
from rest_framework.viewsets import ModelViewSet
from django.views import View
from django.shortcuts import render
from django.contrib import messages
from openpyxl import load_workbook


class CardOperationAddView(View):
    def get(self, request):
        return render(request, 'core/card_operations_add.html')

    def post(self, request):
        excel_file = request.FILES["excel_file"]
        new_excel_source = ExcelSource.objects.create(
            excel_file=excel_file,
            created_by=request.user
        )
        messages.success(request, "Файл добавлен")
        context = {}
        excel_file_source = load_workbook('media/' + new_excel_source.excel_file.name)
        page = excel_file_source[excel_file_source.sheetnames[1]]
        created_qty = 0
        no_count = False
        for row in page:
            if not no_count:
                if row[0].row == 1:
                    no_count = True
                continue
            date = row[0].value
            station = row[1].value
            card = row[2].value
            fuel_type = row[3].value
            operation_type = row[4].value
            balance_before = row[5].value
            balance_after = row[6].value
            dose = row[7].value
            price_som = row[8].value
            sum_som = row[9].value
            new_card_operation, created = CardOperation.objects.get_or_create(
                date=date,
                station=station,
                card=card,
                fuel_type=fuel_type,
                operation_type=operation_type,
                balance_before=balance_before,
                balance_after=balance_after,
                dose=dose,
                price_som=price_som,
                sum_som=sum_som,
            )
            if created:
                created_qty += 1
        messages.success(request, f"Добавлено {created_qty} операции по картам")
        return render(request, 'core/card_operations_add.html', context)

class CardOperationViewSet(ModelViewSet):
    queryset = CardOperation.objects.all()
    serializer_class = CardOperationSerializer